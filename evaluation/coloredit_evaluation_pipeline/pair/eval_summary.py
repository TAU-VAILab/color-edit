

def run_eval_summary_pair(evaluation_results_path,output_path, with_swap=False):
    import os
    import pandas as pd
    import json

    with open(evaluation_results_path, "r") as f:
        data = json.load(f)
    data_folder = data

    # Function to process data
    results = []
    problematic_items = []

    for index, item in enumerate(data_folder):
        index_folder_path = item.get("index_folder_path", "")
        source_visual = item.get("source_visual", "")
        edited_visual = item.get("edited_visual", "")
        
        if not with_swap:
            if not os.path.exists(source_visual) or not os.path.exists(edited_visual):
                # Record problematic items
                problematic_items.append(index_folder_path)
                continue

        ref_color = item.get("ref_color", "").lower()
        color1_name = item.get("color1_name", "").lower()
        color2_name = item.get("color2_name", "").lower()

        if ref_color == color1_name:
            lab_distance_score_source = item.get("source_best_color_distance1", None)
            rgb_l2_score_source = item.get("source_rgb_l2_diff1", None)
            # edit
            lab_distance_score_edited = item.get("edited_best_color_distance1", None)
            rgb_l2_score_edited = item.get("edited_rgb_l2_diff1", None)
            # if any None: print

        elif ref_color == color2_name:
            lab_distance_score_source = item.get("source_best_color_distance2", None)
            rgb_l2_score_source = item.get("source_rgb_l2_diff2", None)
            # edit
            lab_distance_score_edited = item.get("edited_best_color_distance2", None)
            rgb_l2_score_edited = item.get("edited_rgb_l2_diff2", None)
        else:
            # Skip if ref_color does not match
            continue

        results.append({
            "set": item.get("type", ""),
            "method": item.get("method", ""),
            "index": item.get("index", ""),
            "lab_distance_score": lab_distance_score_source,
            "rgb_l2_score": rgb_l2_score_source,
            "ref_color": ref_color,
            "type": "source"
        })
        results.append({
            "set": item.get("type", ""),
            "method": item.get("method", ""),
            "index": item.get("index", ""),
            "lab_distance_score": lab_distance_score_edited,
            "rgb_l2_score": rgb_l2_score_edited,
            "ref_color": ref_color,
            "type": "edited"
        })

    # Create a DataFrame
    df_results = pd.DataFrame(results)

    # # Display results for review
    # import ace_tools as tools; tools.display_dataframe_to_user(name="Evaluation DataFrame", dataframe=df_results)

    # # Save problematic items for reference
    # problematic_items_file = "problematic_items.txt"
    # with open(problematic_items_file, "w") as f:
    #     f.write("\n".join(problematic_items))

    # df_results
    # whole
    # Comprehensive table combining all metrics: average differences, absolute averages, medians, and accuracy
    threshold = 10
    comprehensive_results = []

    # Group by `set` and `method`
    for (set_name, method_name), group in df_results.groupby(['set', 'method']):
        # Separate source and edited rows
        edited_rows = group[group['type'] == 'edited']
        source_rows = group[group['type'] == 'source']
        
        # Compute metrics for source
        source_lab_median = source_rows['lab_distance_score'].median()
        source_rgb_median = source_rows['rgb_l2_score'].median()
        source_lab_mean = source_rows['lab_distance_score'].mean()
        source_rgb_mean = source_rows['rgb_l2_score'].mean()
        source_acc = (source_rows['lab_distance_score'] < threshold).sum() / len(source_rows) if len(source_rows) > 0 else 0

        # Compute metrics for edited
        edited_lab_median = edited_rows['lab_distance_score'].median()
        edited_rgb_median = edited_rows['rgb_l2_score'].median()
        edited_lab_mean = edited_rows['lab_distance_score'].mean()
        edited_rgb_mean = edited_rows['rgb_l2_score'].mean()
        edited_acc = (edited_rows['lab_distance_score'] < threshold).sum() / len(edited_rows) if len(edited_rows) > 0 else 0

        # Compute differences
        lab_diff_mean = edited_lab_mean - source_lab_mean
        rgb_diff_mean = edited_rgb_mean - source_rgb_mean
        lab_diff_median = edited_lab_median - source_lab_median
        rgb_diff_median = edited_rgb_median - source_rgb_median
        acc_diff = edited_acc - source_acc

        # Append results
        comprehensive_results.append({
            'set': set_name,
            'method': method_name,
            'source_lab_mean': source_lab_mean,
            'edited_lab_mean': edited_lab_mean,
            'lab_diff_mean': lab_diff_mean,
            'source_rgb_mean': source_rgb_mean,
            'edited_rgb_mean': edited_rgb_mean,
            'rgb_diff_mean': rgb_diff_mean,
            'source_lab_median': source_lab_median,
            'edited_lab_median': edited_lab_median,
            'lab_diff_median': lab_diff_median,
            'source_rgb_median': source_rgb_median,
            'edited_rgb_median': edited_rgb_median,
            'rgb_diff_median': rgb_diff_median,
            'source_acc': source_acc,
            'edited_acc': edited_acc,
            'acc_diff': acc_diff
        })

    # Create a comprehensive DataFrame
    df_comprehensive = pd.DataFrame(comprehensive_results)

    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font

    # Create an Excel file with the specified structure
    def create_excel_table(file_path, df):
        # Initialize the workbook
        wb = Workbook()

        # Create one sheet to cover both "Close" and "Distant" sets
        ws = wb.active
        ws.title = "Results"

        # Merge cells for headers and subheaders
        ws.merge_cells("A1:A3")  # "Method"
        ws["A1"] = "Method"
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A1"].font = Font(bold=True)

        ws.merge_cells("B1:F1")  # "Close"
        ws["B1"] = "Close"
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["B1"].font = Font(bold=True)

        ws.merge_cells("G1:K1")  # "Distant"
        ws["G1"] = "Distant"
        ws["G1"].alignment = Alignment(horizontal="center", vertical="center")
        ws["G1"].font = Font(bold=True)

        # Add subheaders for Lab, RGB_L2, and Accuracy under Close and Distant
        headers = ["Lab", "RGB_L2", "Accuracy", "Lab", "RGB_L2", "Accuracy"]
        col_start = 2
        for i, header in enumerate(headers):
            if header != "Accuracy":
                ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_start+1)
                ws.cell(row=2, column=col_start).value = header
                ws.cell(row=2, column=col_start).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=2, column=col_start).font = Font(bold=True)
                col_start += 2
            else:
                ws.cell(row=2, column=col_start).value = header
                ws.cell(row=2, column=col_start).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=2, column=col_start).font = Font(bold=True)
                col_start += 1

        # Add metric labels: Mean, Median
        col_start = 2
        for _ in range(2):  # Loop twice: once for Close, once for Distant
            for label in ["Mean", "Median", "Mean", "Median", "Mean"]:
                ws.cell(row=3, column=col_start).value = label
                ws.cell(row=3, column=col_start).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=3, column=col_start).font = Font(bold=True)
                col_start += 1

        # Add rows for each method and Source/Edited
        row_start = 4
        for method in df["method"].unique():
            # Add method name
            ws.cell(row=row_start, column=1, value=method)
            ws.cell(row=row_start, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row_start + 1, column=1, value="")  # Leave the next row empty for alignment

            # Fill in Source and Edited rows for each metric
            for set_type, col_offset in [("close", 2), ("distant", 7)]:
                for row_offset, source_or_edited in enumerate(["source", "edited"]):
                    row = row_start + row_offset
                    for col_idx, metric in enumerate(
                        [f"{source_or_edited}_lab_mean", f"{source_or_edited}_lab_median",
                        f"{source_or_edited}_rgb_mean", f"{source_or_edited}_rgb_median",
                        f"{source_or_edited}_acc"]
                    ):
                        value = df.loc[(df["method"] == method) & (df["set"] == set_type), metric].values
                        if len(value) > 0:
                            ws.cell(row=row, column=col_offset + col_idx, value=value[0])

                    # Add "Source" or "Edited" label
                    if col_offset == 2 and col_idx == 0:  # Only add labels for "Close"
                        ws.cell(row=row, column=1).value = "Source" if source_or_edited == "source" else "Edited"

            row_start += 2  # Two rows per method (Source and Edited)

        # Adjust column widths for better visibility
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        # Save the workbook
        wb.save(file_path)


    # Example usage
    file_path= os.path.join(output_path, "final_corrected_table.xlsx")
    df = df_comprehensive
    create_excel_table(file_path, df)
